"""
Check Multiple Presenter - Handles batch operations on multiple connectors
"""
from PySide6.QtCore import QObject, Signal
from PySide6.QtWidgets import QFileDialog
from .view import CheckMultipleConnectorView
from .config import OPERATION_RESULT_COLUMNS
from ...e3 import E3Model
import pandas as pd
from pathlib import Path


class CheckMultipleConnectorPresenter(QObject):
    """Presenter for Check Multiple functionality"""

    # Signal to request tab switch with search text
    switch_to_lookup = Signal(str)  # part_numbers_str

    def __init__(self, context, connector_model):
        super().__init__()
        self.context = context
        self.model = connector_model
        self.view = CheckMultipleConnectorView()

        # Initialize E3 model
        self.e3_model = E3Model(self.context)

        # Data storage
        self.imported_df = None
        self.search_column = None
        self.context_columns = []

        # Connect view signals
        self.view.file_imported.connect(self.on_file_imported)
        self.view.operation_requested.connect(self.on_operation_requested)
        self.view.export_requested.connect(self.on_export)
        self.view.clear_results_requested.connect(self.on_clear_results)
        self.view.remove_data_requested.connect(self.on_remove_data)
        self.view.to_lookup_requested.connect(self.on_to_lookup_requested)
        self.view.group_by_requested.connect(self.on_group_by_requested)
        self.view.file_dialog_shown.connect(self.setup_file_dialog_connections)

        # Connect E3 model signals
        self.e3_model.loading_progress.connect(self._on_e3_progress)
        self.e3_model.loading_failed.connect(self._on_e3_error)
        self.e3_model.projects_loaded.connect(self._on_e3_projects_loaded)
        self.e3_model.connectors_loaded.connect(self._on_e3_connectors_loaded)

    def start_loading(self):
        """Placeholder for loading"""
        pass

    def on_file_imported(self, file_path: str, search_column: str, context_columns: list):
        """Handle file import"""
        print(
            f"File imported: {file_path}, search column: {search_column}, context columns: {context_columns}")

        try:
            # Load the file
            path = Path(file_path)

            if path.suffix.lower() == '.csv':
                self.imported_df = pd.read_csv(file_path)
            elif path.suffix.lower() == '.xlsx':
                self.imported_df = pd.read_excel(file_path)
            elif path.suffix.lower() == '.txt':
                self.imported_df = pd.read_csv(
                    file_path, sep=None, engine='python')

            self.search_column = search_column
            self.context_columns = context_columns

            # Rename columns to indicate they are user inputs
            # Prefix all columns with "Input: "
            columns_to_drop = [col for col in self.imported_df.columns if col not in [
                self.search_column] + self.context_columns]
            self.imported_df.drop(columns=columns_to_drop, inplace=True)

            renamed_columns = {
                col: f"Input: {col}" for col in self.imported_df.columns}
            self.imported_df.rename(columns=renamed_columns, inplace=True)

            # Update the search_column and context_columns references
            self.search_column = f"Input: {search_column}"
            self.context_columns = [f"Input: {col}" for col in context_columns]

            # Reorder columns to ensure input columns are first (search column first, then context columns, then others)
            self.imported_df = self._reorder_imported_columns(self.imported_df)

            print(f"Loaded {len(self.imported_df)} rows")

            # Show the imported data in the results table (mark as original import)
            self.view.update_results(self.imported_df, is_original_import=True)

        except Exception as e:
            self.view.show_error(f"Failed to import file: {str(e)}")
            print(f"Error importing file: {e}")

    def _reorder_imported_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Reorder imported columns: search column first, then context columns, then others"""
        if df.empty:
            return df

        # Get all columns
        all_cols = list(df.columns)

        # Build new order: search column, context columns, then remaining
        new_order = []

        # Add search column first
        if self.search_column in all_cols:
            new_order.append(self.search_column)

        # Add context columns next
        for col in self.context_columns:
            if col in all_cols and col not in new_order:
                new_order.append(col)

        # Add remaining columns
        for col in all_cols:
            if col not in new_order:
                new_order.append(col)

        return df[new_order]

    def on_operation_requested(self, operation_type: str, filters: dict):
        """Handle batch operation request"""
        print(f"Operation requested: {operation_type}, filters: {filters}")

        if self.imported_df is None:
            self.view.show_error("No file imported")
            return

        self.view.show_loading(True)

        # Get search terms from the search column
        search_terms = self.imported_df[self.search_column].dropna(
        ).unique().tolist()

        print(f"Processing {len(search_terms)} unique search terms")

        # Process based on operation type
        if operation_type == 'find_opposites':
            results = self._batch_find_opposites(search_terms)
        elif operation_type == 'find_alternatives':
            results = self._batch_find_alternatives(search_terms, filters)
        elif operation_type == 'lookup':
            results = self._batch_lookup(search_terms)
        elif operation_type == 'get_material':
            results = self._batch_get_material(search_terms)
        elif operation_type == 'check_status':
            results = self._batch_check_status(search_terms)
        else:
            results = pd.DataFrame()

        # Filter columns based on operation type
        if results is not None and not results.empty:
            results = self._filter_result_columns(results, operation_type)

        # Add context columns if any
        if results is not None and not results.empty:
            results = self._merge_context_columns(results)
            # Ensure column order: Input columns first, then Status, Search Term, then other result columns
            results = self._reorder_columns(results)

        self.view.show_loading(False)

        if results is not None and not results.empty:
            self.view.update_results(results)
        else:
            self.view.show_error("No results found")

    def _batch_find_opposites(self, search_terms: list) -> pd.DataFrame:
        """Find opposite connectors for all search terms"""
        results = []

        for term in search_terms:
            # Initialize result record
            result_record = {'Search Term': term, 'Status': 'Not Found'}

            if hasattr(self.model, 'find_opposite'):
                opposite = self.model.find_opposite(term)
                if opposite:
                    result_record['Status'] = 'Found'
                    # Merge with opposite data
                    result_record.update(opposite)

            results.append(result_record)

        return pd.DataFrame(results) if results else pd.DataFrame()

    def _batch_find_alternatives(self, search_terms: list, filters: dict) -> pd.DataFrame:
        """Find alternative connectors for all search terms"""
        results = []

        for term in search_terms:
            if hasattr(self.model, 'find_alternative'):
                alternatives = self.model.find_alternative(term)
                if alternatives:
                    for alt in alternatives:
                        result_record = {
                            'Search Term': term, 'Status': 'Found'}
                        result_record.update(alt)
                        results.append(result_record)
                else:
                    # No alternatives found
                    results.append(
                        {'Search Term': term, 'Status': 'Not Found'})
            else:
                results.append({'Search Term': term, 'Status': 'Not Found'})

        df = pd.DataFrame(results) if results else pd.DataFrame()

        # Apply filters if provided
        if not df.empty and filters:
            for filter_key, filter_value in filters.items():
                # Map filter keys to column names
                column_mapping = {
                    'standard': 'Family',
                    'shell_type': 'Shell Type',
                    'material': 'Material',
                    'shell_size': 'Shell Size',
                    'insert_arrangement': 'Insert Arrangement',
                    'socket_type': 'Socket Type',
                    'keying': 'Keying'
                }

                column_name = column_mapping.get(filter_key)
                if column_name and column_name in df.columns:
                    # Handle "Same" filter - need to filter based on original search term
                    if filter_value == "Same":
                        # This would require looking up the original value for each search term
                        # For now, skip "Same" filter in batch mode
                        continue
                    df = df[df[column_name] == filter_value]

        return df

    def _batch_lookup(self, search_terms: list) -> pd.DataFrame:
        """Lookup connectors for all search terms - returns ALL connector details"""
        print(f"Batch lookup for {len(search_terms)} terms")
        results = []

        for term in search_terms:
            print(f"DEBUG: Searching for term: '{term}'")
            # Use the model to search for the connector
            matches = self.model.filter_connectors({'search_text': str(term)})
            print(f"DEBUG: Found {len(matches)} matches for '{term}'")

            if matches:
                # Add all matching connectors with full details
                for match in matches:
                    result_record = {'Search Term': term, 'Status': 'Found'}
                    # Add ALL fields from the connector
                    result_record.update(match)
                    results.append(result_record)
                    print(
                        f"DEBUG: Added result with keys: {list(result_record.keys())}")
            else:
                # No matches found - add empty row with just Search Term and Status
                results.append({'Search Term': term, 'Status': 'Not Found'})
                print(f"DEBUG: No matches for '{term}'")

        df = pd.DataFrame(results) if results else pd.DataFrame()
        if not df.empty:
            print(f"DEBUG: Result DataFrame columns: {df.columns.tolist()}")
            print(f"DEBUG: Result DataFrame shape: {df.shape}")
        return df

    def _batch_get_material(self, search_terms: list) -> pd.DataFrame:
        """Get material information for all search terms - returns ALL connector details"""
        print(f"Batch get material for {len(search_terms)} terms")
        results = []

        for term in search_terms:
            # Use the model to search for the connector
            matches = self.model.filter_connectors({'search_text': str(term)})

            if matches:
                # Add all matching connectors with full details
                for match in matches:
                    result_record = {'Search Term': term, 'Status': 'Found'}
                    # Add ALL fields from the connector
                    result_record.update(match)
                    results.append(result_record)
            else:
                # No matches found
                results.append({'Search Term': term, 'Status': 'Not Found'})

        return pd.DataFrame(results) if results else pd.DataFrame()

    def _batch_check_status(self, search_terms: list) -> pd.DataFrame:
        """Check database status for all search terms - returns ALL connector details"""
        print(f"Batch check status for {len(search_terms)} terms")
        results = []

        for term in search_terms:
            # Use the model to search for the connector
            matches = self.model.filter_connectors({'search_text': str(term)})

            if matches:
                # Add all matching connectors with full details
                for match in matches:
                    result_record = {'Search Term': term, 'Status': 'Found'}
                    # Add ALL fields from the connector
                    result_record.update(match)
                    results.append(result_record)
            else:
                # No matches found
                results.append({'Search Term': term, 'Status': 'Not Found'})

        return pd.DataFrame(results) if results else pd.DataFrame()

    def _filter_result_columns(self, results_df: pd.DataFrame, operation_type: str) -> pd.DataFrame:
        """Filter result columns based on operation configuration

        Args:
            results_df: DataFrame with all result columns
            operation_type: Type of operation (lookup, get_material, check_status, etc.)

        Returns:
            DataFrame with filtered columns
        """
        if results_df.empty:
            return results_df

        # Get column configuration for this operation
        column_config = OPERATION_RESULT_COLUMNS.get(operation_type, 'all')

        # If 'all', return everything
        if column_config == 'all':
            print(
                f"DEBUG: Operation '{operation_type}' configured to show all columns")
            return results_df

        # Otherwise, keep only specified columns (plus Search Term and Status which are always needed)
        columns_to_keep = ['Search Term', 'Status']

        # Add configured columns that exist in the dataframe
        for col in column_config:
            if col in results_df.columns:
                columns_to_keep.append(col)

        print(
            f"DEBUG: Filtering '{operation_type}' to columns: {columns_to_keep}")

        # Filter the dataframe
        filtered_df = results_df[columns_to_keep].copy()

        return filtered_df

    def _merge_context_columns(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """Merge ONLY selected input columns from imported data with results"""
        if self.imported_df is None:
            return results_df

        # Merge on Search Term if available
        if 'Search Term' in results_df.columns and self.search_column in self.imported_df.columns:
            # Get list of columns that will cause duplicates (exist in both dataframes)
            # We want to keep the input version of these, not the result version
            input_cols = set(self.imported_df.columns)
            result_cols = set(results_df.columns)
            overlapping_cols = input_cols & result_cols

            # Remove overlapping columns from results BEFORE merge (except Search Term which we need for merging)
            cols_to_drop_from_results = [
                col for col in overlapping_cols if col != 'Search Term']
            if cols_to_drop_from_results:
                results_df = results_df.drop(columns=cols_to_drop_from_results)

            # Create a copy of imported data for merging - ONLY search column and selected context columns
            columns_to_merge = [self.search_column] + self.context_columns
            # Only include columns that actually exist
            columns_to_merge = [
                col for col in columns_to_merge if col in self.imported_df.columns]

            print(
                f"DEBUG: Merging only these input columns: {columns_to_merge}")

            input_df = self.imported_df[columns_to_merge].copy()

            # Add a temporary merge key that matches Search Term values
            # But DON'T rename the original search column - keep it!
            input_df['_temp_merge_key'] = input_df[self.search_column]

            # Remove duplicates based on the merge key (in case multiple rows have same search value)
            input_df = input_df.drop_duplicates(subset=['_temp_merge_key'])

            # Merge using the temporary key
            merged = results_df.merge(
                input_df,
                left_on='Search Term',
                right_on='_temp_merge_key',
                how='left'
            )

            # Remove the temporary merge key AND the Search Term column
            # We keep the original search column (e.g., "Input: Part Number") instead
            cols_to_remove = ['_temp_merge_key', 'Search Term']
            merged = merged.drop(
                columns=[col for col in cols_to_remove if col in merged.columns])

            return merged

        return results_df

    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Reorder columns to ensure Input columns appear first"""
        if df.empty:
            return df

        # Separate columns into categories
        # Order input columns: search column first, then context columns, then any other input columns
        search_col = [
            self.search_column] if self.search_column in df.columns else []
        context_cols = [
            col for col in self.context_columns if col in df.columns]
        other_input_cols = [col for col in df.columns
                            if col.startswith('Input: ')
                            and col not in search_col
                            and col not in context_cols]

        # All input columns in order
        input_cols = search_col + context_cols + other_input_cols

        # Status column
        status_col = ['Status'] if 'Status' in df.columns else []

        # All other result columns (excluding input columns and status)
        other_cols = [col for col in df.columns
                      if col not in input_cols
                      and col not in status_col]

        # New order: Input columns (search, context, others), Status, then other result columns
        new_order = input_cols + status_col + other_cols

        # Only include columns that exist in the dataframe
        new_order = [col for col in new_order if col in df.columns]

        return df[new_order]

    def on_export(self):
        """Handle export request"""
        print("Export requested")

        if self.view.imported_data is None:
            self.view.show_error("No data to export")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Export Results",
            "",
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)"
        )

        if file_path:
            try:
                path = Path(file_path)

                # Check if we have grouped data
                if self.view.current_grouping:
                    print(
                        f"Exporting grouped data by {self.view.current_grouping}")
                    self._export_grouped(file_path, self.view.current_grouping)
                else:
                    print("Exporting flat data")
                    # Export flat data as before
                    if path.suffix.lower() == '.csv':
                        self.view.imported_data.to_csv(file_path, index=False)
                    elif path.suffix.lower() in ['.xlsx', '.xls']:
                        self.view.imported_data.to_excel(
                            file_path, index=False)
                    else:
                        # Default to CSV
                        self.view.imported_data.to_csv(file_path, index=False)

                print(f"Exported to {file_path}")
                self.view.record_count_label.setText(
                    f"Exported {len(self.view.imported_data)} rows to {path.name}")
            except Exception as e:
                self.view.show_error(f"Export failed: {str(e)}")
                print(f"Export error: {e}")

    def _export_grouped(self, file_path: str, group_field: str):
        """Export data with grouped structure preserved"""
        path = Path(file_path)
        df = self.view.imported_data

        # Group the data
        grouped = df.groupby(group_field, sort=False)

        # Sort groups by size (descending)
        group_sizes = grouped.size().sort_values(ascending=False)

        # Create export data with group headers
        export_rows = []

        for group_value in group_sizes.index:
            group_data = grouped.get_group(group_value)
            group_count = len(group_data)

            # Create group header row
            group_header = [f"{group_field}: {group_value} ({group_count})"]
            # Empty cells for remaining columns
            group_header.extend([''] * (len(df.columns) - 1))
            export_rows.append(group_header)

            # Add group data rows
            for _, row in group_data.iterrows():
                export_rows.append(row.tolist())

        # Create DataFrame for export
        export_df = pd.DataFrame(export_rows, columns=df.columns)

        # Export based on file type
        if path.suffix.lower() == '.csv':
            export_df.to_csv(file_path, index=False)
        elif path.suffix.lower() in ['.xlsx', '.xls']:
            # For Excel, we can make it fancier with formatting
            self._export_grouped_excel(
                file_path, group_field, grouped, group_sizes)
        else:
            # Default to CSV
            export_df.to_csv(file_path, index=False)

    def _export_grouped_excel(self, file_path: str, group_field: str, grouped, group_sizes):
        """Export grouped data to Excel with formatting"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # If openpyxl not available, fall back to basic export
            print("openpyxl not available, using basic Excel export")
            df = self.view.imported_data
            df.to_excel(file_path, index=False)
            return

        df = self.view.imported_data
        wb = Workbook()
        ws = wb.active
        ws.title = "Grouped Results"

        # Write header row
        ws.append(df.columns.tolist())

        # Style for header
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Style for group headers
        group_font = Font(bold=True, color="FFFFFF")
        group_fill = PatternFill(start_color="6495ED",
                                 end_color="6495ED", fill_type="solid")

        # Write data by groups
        for group_value in group_sizes.index:
            group_data = grouped.get_group(group_value)
            group_count = len(group_data)

            # Add group header row
            group_header_row = [
                f"{group_field}: {group_value} ({group_count})"]
            group_header_row.extend([''] * (len(df.columns) - 1))
            ws.append(group_header_row)

            # Style the group header row
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.font = group_font
                cell.fill = group_fill
                cell.alignment = Alignment(horizontal='left')

            # Add group data rows
            for _, row in group_data.iterrows():
                ws.append(row.tolist())

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def on_clear_results(self):
        """Handle clear results request - handled by view directly"""
        pass

    def on_remove_data(self):
        """Handle remove data request"""
        # Clear presenter data
        self.imported_df = None
        self.search_column = None
        self.context_columns = []
        print("Data removed")

    def on_to_lookup_requested(self, search_terms: list):
        """Handle request to switch to Lookup tab with part numbers

        Args:
            search_terms: List of search terms from the imported data
                         (might not be actual part numbers yet, OR could be Part Numbers from context menu)
        """
        print(
            f"DEBUG PRESENTER: on_to_lookup_requested called with {len(search_terms)} search terms")
        # Show first 5
        print(f"DEBUG PRESENTER: Search terms: {search_terms[:5]}...")

        # Check if these look like actual Part Numbers (context menu sends Part Numbers directly)
        # Part Numbers typically contain specific patterns like slashes, dashes, or specific prefixes
        likely_part_numbers = all(
            any(indicator in str(term)
                for indicator in ['/', '-', 'MS', 'EN', 'VG', 'MIL', 'D38999'])
            for term in search_terms[:5]  # Check first 5
        )

        if likely_part_numbers:
            print("DEBUG PRESENTER: Terms appear to be Part Numbers already, using as-is")
            part_numbers_list = [str(term) for term in search_terms]
        else:
            print("DEBUG PRESENTER: Terms appear to be search terms, performing lookup")
            # Perform lookup for each search term to get actual Part Numbers
            part_numbers = set()  # Use set to avoid duplicates

            for term in search_terms:
                print(f"DEBUG PRESENTER: Looking up term: {term}")

                # Try multiple search strategies
                # Strategy 1: Search with filter_connectors
                results = self.model.filter_connectors(
                    {'search_text': str(term)})
                print(
                    f"DEBUG PRESENTER: filter_connectors returned {len(results)} results")

                # Strategy 2: If no results, try getting all connectors and filter manually
                if not results:
                    all_connectors = self.model.get_connectors()
                    print(
                        f"DEBUG PRESENTER: Got {len(all_connectors)} total connectors, filtering manually")

                    # Search in Part Number and Part Code fields
                    term_lower = str(term).lower()
                    for conn in all_connectors:
                        part_num = str(conn.get('Part Number', '')).lower()
                        part_code = str(conn.get('Part Code', '')).lower()

                        if term_lower in part_num or term_lower in part_code:
                            results.append(conn)

                    print(
                        f"DEBUG PRESENTER: Manual filter found {len(results)} results")

                # Extract Part Numbers from results
                for result in results:
                    part_number = result.get('Part Number')
                    if part_number:
                        part_numbers.add(part_number)
                        print(
                            f"DEBUG PRESENTER: Added part number: {part_number}")

            if not part_numbers:
                print("DEBUG PRESENTER: No part numbers found from lookup")
                # If still no results, just use the search terms as-is
                # They might already be part numbers
                print("DEBUG PRESENTER: Using search terms as part numbers")
                part_numbers = set(str(term) for term in search_terms)

            # Convert to sorted list
            part_numbers_list = sorted(list(part_numbers))

        # Create comma-separated list of actual Part Numbers
        part_numbers_str = ", ".join(part_numbers_list)
        print(
            f"DEBUG PRESENTER: Found {len(part_numbers_list)} part numbers: {part_numbers_str[:100]}...")

        # Emit signal that will be handled by the connector module
        self.switch_to_lookup.emit(part_numbers_str)
        print("DEBUG PRESENTER: switch_to_lookup signal emitted")

    def on_group_by_requested(self, field: str):
        """Handle group by request from view"""
        print(f"DEBUG PRESENTER: Group by requested for field: {field}")

        # Check if we have data to group
        if self.view.imported_data is None or self.view.imported_data.empty:
            print("DEBUG PRESENTER: No data to group")
            return

        print(
            f"DEBUG PRESENTER: Available columns: {self.view.imported_data.columns.tolist()}")
        print(f"DEBUG PRESENTER: Data shape: {self.view.imported_data.shape}")

        # Check if the field exists in the data
        if field not in self.view.imported_data.columns:
            print(
                f"DEBUG PRESENTER: Field '{field}' not found in data columns")
            print(f"DEBUG PRESENTER: Looking for: '{field}'")
            print(
                f"DEBUG PRESENTER: Available: {self.view.imported_data.columns.tolist()}")
            return

        # Call view's update_grouped_results
        self.view.update_grouped_results(self.view.imported_data, field)
        print(f"DEBUG PRESENTER: Called update_grouped_results for {field}")

    def setup_file_dialog_connections(self, dialog):
        """Connect file dialog E3 signals to presenter handlers

        This should be called when the file upload dialog is shown

        Args:
            dialog: FileUploadDialog instance
        """
        # Connect dialog signals to presenter
        dialog.request_e3_projects_available.connect(
            self._on_request_e3_projects)
        dialog.request_e3_project_caches_available.connect(
            self._on_request_e3_caches)

        # Connect E3 model signals to dialog
        self.e3_model.projects_loaded.connect(dialog.populate_e3_projects)

    def _on_request_e3_projects(self):
        """Handle request for E3 projects"""
        print("E3: Loading available projects...")
        self.e3_model.load_available_projects_async()

    def _on_request_e3_caches(self):
        """Handle request for E3 cache files"""
        print("E3: Loading available cache files...")
        caches = self.e3_model.get_available_cache_files()
        # NOTE: We need to emit this to the dialog, but we don't have direct reference
        # The dialog will be connected via setup_file_dialog_connections
        print(f"E3: Found {len(caches)} cache files: {caches}")

    def _on_e3_progress(self, percent: int, message: str):
        """Handle E3 loading progress updates"""
        print(f"E3 Progress ({percent}%): {message}")
        # TODO: Show progress in UI (progress bar in dialog)

    def _on_e3_error(self, error_message: str):
        """Handle E3 loading errors"""
        print(f"E3 Error: {error_message}")
        # TODO: Show error dialog to user

    def _on_e3_projects_loaded(self, projects: list):
        """Handle E3 projects loaded successfully"""
        print(f"E3: {len(projects)} projects loaded: {projects}")
        # Signal is already connected to dialog.populate_e3_projects

    def _on_e3_connectors_loaded(self, data):
        """Handle E3 connectors loaded

        Args:
            data: DataFrame containing connector data from E3
        """
        print(f"E3: Connectors loaded: {len(data)} rows")

        # Treat E3 data as if it was an imported file
        # Set it as imported_df and trigger normal workflow
        self.imported_df = data

        # Add "Input: " prefix to column names to match file import behavior
        renamed_columns = {col: f"Input: {col}" for col in data.columns}
        self.imported_df.rename(columns=renamed_columns, inplace=True)

        # Set search column to the first column (typically Part Number)
        if len(self.imported_df.columns) > 0:
            self.search_column = self.imported_df.columns[0]
            # All other columns are context columns
            self.context_columns = [
                col for col in self.imported_df.columns if col != self.search_column]

            print(f"E3: Search column: {self.search_column}")
            print(f"E3: Context columns: {self.context_columns}")

            # Display in results table
            self.view.display_table(self.imported_df, grouped=False)
            print(
                f"E3: Displayed {len(self.imported_df)} rows from E3 projects")
